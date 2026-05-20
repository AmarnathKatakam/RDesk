"""
payslip_generation.report_views — Milestone 4: Payroll Reports

Endpoints:
  GET /api/payroll/reports/register/          — Payroll Register (all employees, all components)
  GET /api/payroll/reports/register/export/   — Export register as Excel (.xlsx)
  GET /api/payroll/reports/bank-transfer/     — Bank Transfer file data
  GET /api/payroll/reports/bank-transfer/export/ — Export bank transfer as CSV
  GET /api/payroll/reports/department-summary/ — Department-wise totals
  GET /api/payroll/reports/variance/          — Month-over-month variance per employee

All endpoints accept query params: month, year, salary_type (default SALARY)
Register and bank-transfer also accept: run_id (optional — use a specific run)
"""
from __future__ import annotations

import csv
import io
import logging
from decimal import Decimal
from typing import Optional

from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from employees.models import Employee, MonthlySalaryData
from .models import PayrollRun, PayrollRunItem, PayrollRunItemLine

logger = logging.getLogger('payroll.reports')

ZERO = Decimal('0')

MONTH_NUMBER = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
}


def _prev_month(month: str, year: int) -> tuple[str, int]:
    """Return (prev_month_name, prev_year)."""
    mn = MONTH_NUMBER.get(month.lower(), 1)
    if mn == 1:
        return 'December', year - 1
    months = list(MONTH_NUMBER.keys())
    return months[mn - 2].capitalize(), year


def _get_run(month: str, year: int, salary_type: str) -> Optional[PayrollRun]:
    return PayrollRun.objects.filter(
        month=month, year=year, salary_type=salary_type
    ).first()


def _items_for_run(run: PayrollRun):
    return (
        PayrollRunItem.objects
        .filter(run=run, status='INCLUDED')
        .select_related('employee', 'employee__department')
        .prefetch_related('lines')
        .order_by('employee__name')
    )


def _build_register_row(item: PayrollRunItem) -> dict:
    """Build one register row from a PayrollRunItem (3C engine or legacy)."""
    emp = item.employee
    lines = list(item.lines.all())

    # Build component map from lines
    comp_map: dict[str, Decimal] = {}
    for line in lines:
        comp_map[line.code] = line.amount

    # Statutory deductions
    pf_emp = comp_map.get('PF_EMP', ZERO)
    esi_emp = comp_map.get('ESI_EMP', ZERO)
    pt = comp_map.get('PT', ZERO)
    lwf_emp = comp_map.get('LWF_EMP', ZERO)
    pf_employer = comp_map.get('PF_EMPLOYER', ZERO)
    esi_employer = comp_map.get('ESI_EMPLOYER', ZERO)

    # Earnings from lines
    earnings = {
        line.code: float(line.amount)
        for line in lines if line.component_type == 'EARNING'
    }
    deductions = {
        line.code: float(line.amount)
        for line in lines if line.component_type == 'DEDUCTION'
    }

    return {
        'employee_pk': emp.id,
        'employee_id': emp.employee_id,
        'employee_name': emp.name,
        'department': emp.department.department_name if emp.department else '',
        'position': emp.position,
        'bank_account': emp.bank_account or '',
        'bank_ifsc': emp.bank_ifsc or '',
        'pan': emp.pan or '',
        'pf_number': emp.pf_number or '',
        'pay_mode': emp.pay_mode,
        'lop_days': item.lop_days,
        'work_days': item.work_days,
        'payable_days': item.payable_days,
        'days_in_month': item.days_in_month,
        'proration_factor': float(item.proration_factor),
        'gross_earnings': float(item.gross_earnings),
        'total_deductions': float(item.total_deductions),
        'employer_contributions': float(item.employer_contributions),
        'net_pay': float(item.net_pay),
        'pf_employee': float(pf_emp),
        'esi_employee': float(esi_emp),
        'professional_tax': float(pt),
        'lwf_employee': float(lwf_emp),
        'pf_employer': float(pf_employer),
        'esi_employer': float(esi_employer),
        'calculation_source': item.calculation_source,
        'earnings_breakdown': earnings,
        'deductions_breakdown': deductions,
    }


def _build_register_from_salary_data(month: str, year: int, salary_type: str) -> list[dict]:
    """Fallback: build register from MonthlySalaryData when no PayrollRun exists."""
    qs = MonthlySalaryData.objects.filter(
        month=month, year=year, salary_type=salary_type
    ).select_related('employee', 'employee__department')

    rows = []
    for sd in qs:
        emp = sd.employee
        rows.append({
            'employee_pk': emp.id,
            'employee_id': emp.employee_id,
            'employee_name': emp.name,
            'department': emp.department.department_name if emp.department else '',
            'position': emp.position,
            'bank_account': emp.bank_account or '',
            'bank_ifsc': emp.bank_ifsc or '',
            'pan': emp.pan or '',
            'pf_number': emp.pf_number or '',
            'pay_mode': emp.pay_mode,
            'lop_days': sd.effective_lop,
            'work_days': sd.work_days,
            'payable_days': sd.work_days - sd.effective_lop,
            'days_in_month': sd.days_in_month,
            'proration_factor': 1.0,
            'gross_earnings': float(sd.gross_earnings),
            'total_deductions': float(sd.total_deductions),
            'employer_contributions': float(sd.pf_employer),
            'net_pay': float(sd.net_pay),
            'pf_employee': float(sd.pf_employee),
            'esi_employee': 0.0,
            'professional_tax': float(sd.professional_tax),
            'lwf_employee': 0.0,
            'pf_employer': float(sd.pf_employer),
            'esi_employer': 0.0,
            'calculation_source': 'MONTHLY_SALARY_DATA',
            'earnings_breakdown': {
                'BASIC': float(sd.basic),
                'HRA': float(sd.hra),
                'DA': float(sd.da),
                'CONVEYANCE': float(sd.conveyance),
                'MEDICAL': float(sd.medical),
                'SPECIAL_ALLOWANCE': float(sd.special_allowance),
            },
            'deductions_breakdown': {
                'PF_EMP': float(sd.pf_employee),
                'PT': float(sd.professional_tax),
                'OTHER': float(sd.other_deductions),
                'ADVANCE': float(sd.salary_advance),
            },
        })
    return rows


def _get_register_rows(month: str, year: int, salary_type: str) -> tuple[list[dict], Optional[PayrollRun]]:
    """Return (rows, run_or_None). Prefers PayrollRun data, falls back to MonthlySalaryData."""
    run = _get_run(month, year, salary_type)
    if run:
        items = _items_for_run(run)
        rows = [_build_register_row(item) for item in items]
    else:
        rows = _build_register_from_salary_data(month, year, salary_type)
    return rows, run


# ─── Payroll Register ─────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payroll_register(request):
    """
    GET /api/payroll/reports/register/
    Returns full payroll register for a period.
    Query params: month, year, salary_type (default SALARY)
    """
    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    try:
        year = int(year_str)
    except ValueError:
        return Response({'success': False, 'message': 'year must be an integer.'},
                        status=status.HTTP_400_BAD_REQUEST)

    rows, run = _get_register_rows(month, year, salary_type)

    # Collect all unique earning/deduction codes for column headers
    all_earning_codes = sorted({
        code for row in rows for code in row.get('earnings_breakdown', {})
    })
    all_deduction_codes = sorted({
        code for row in rows for code in row.get('deductions_breakdown', {})
    })

    # Summary totals
    total_gross = sum(r['gross_earnings'] for r in rows)
    total_deductions = sum(r['total_deductions'] for r in rows)
    total_net = sum(r['net_pay'] for r in rows)
    total_employer = sum(r['employer_contributions'] for r in rows)

    return Response({
        'success': True,
        'month': month,
        'year': year,
        'salary_type': salary_type,
        'run_id': run.id if run else None,
        'run_status': run.status if run else None,
        'employee_count': len(rows),
        'summary': {
            'total_gross': round(total_gross, 2),
            'total_deductions': round(total_deductions, 2),
            'total_net': round(total_net, 2),
            'total_employer_contributions': round(total_employer, 2),
        },
        'columns': {
            'earning_codes': all_earning_codes,
            'deduction_codes': all_deduction_codes,
        },
        'rows': rows,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payroll_register_export(request):
    """
    GET /api/payroll/reports/register/export/
    Returns Excel (.xlsx) payroll register.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({'success': False, 'message': 'openpyxl not installed. Run: pip install openpyxl'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    year = int(year_str)

    rows, run = _get_register_rows(month, year, salary_type)

    # Collect all unique codes
    all_earning_codes = sorted({code for row in rows for code in row.get('earnings_breakdown', {})})
    all_deduction_codes = sorted({code for row in rows for code in row.get('deductions_breakdown', {})})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll Register {month} {year}"

    # Header style
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center')

    # Fixed columns
    fixed_cols = [
        'Employee ID', 'Name', 'Department', 'Position', 'PAN', 'PF Number',
        'Bank Account', 'Bank IFSC', 'Pay Mode',
        'Days in Month', 'Work Days', 'LOP Days', 'Payable Days',
    ]
    earning_labels = [f'E: {c}' for c in all_earning_codes]
    deduction_labels = [f'D: {c}' for c in all_deduction_codes]
    summary_cols = [
        'Gross Earnings', 'Total Deductions', 'Employer Contributions', 'Net Pay',
        'PF Employee', 'ESI Employee', 'Prof. Tax', 'PF Employer', 'ESI Employer',
    ]

    all_headers = fixed_cols + earning_labels + deduction_labels + summary_cols

    # Write headers
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        eb = row.get('earnings_breakdown', {})
        db = row.get('deductions_breakdown', {})
        values = [
            row['employee_id'], row['employee_name'], row['department'], row['position'],
            row['pan'], row['pf_number'], row['bank_account'], row['bank_ifsc'], row['pay_mode'],
            row['days_in_month'], row['work_days'], row['lop_days'], row['payable_days'],
        ]
        values += [eb.get(c, 0) for c in all_earning_codes]
        values += [db.get(c, 0) for c in all_deduction_codes]
        values += [
            row['gross_earnings'], row['total_deductions'],
            row['employer_contributions'], row['net_pay'],
            row['pf_employee'], row['esi_employee'], row['professional_tax'],
            row['pf_employer'], row['esi_employer'],
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Totals row
    total_row = row_idx + 2 if rows else 3
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    # Sum numeric columns
    for col_idx in range(len(fixed_cols) + 1, len(all_headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}2:{col_letter}{total_row - 2})').font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"payroll_register_{month}_{year}_{salary_type}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Bank Transfer Report ─────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bank_transfer_report(request):
    """
    GET /api/payroll/reports/bank-transfer/
    Returns bank transfer data: employee, account, IFSC, net pay.
    """
    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    year = int(year_str)

    rows, run = _get_register_rows(month, year, salary_type)

    bank_rows = []
    for row in rows:
        bank_rows.append({
            'employee_id': row['employee_id'],
            'employee_name': row['employee_name'],
            'department': row['department'],
            'bank_account': row['bank_account'],
            'bank_ifsc': row['bank_ifsc'],
            'pay_mode': row['pay_mode'],
            'net_pay': row['net_pay'],
        })

    total_transfer = sum(r['net_pay'] for r in bank_rows)

    return Response({
        'success': True,
        'month': month,
        'year': year,
        'salary_type': salary_type,
        'run_id': run.id if run else None,
        'employee_count': len(bank_rows),
        'total_transfer_amount': round(total_transfer, 2),
        'rows': bank_rows,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bank_transfer_export(request):
    """
    GET /api/payroll/reports/bank-transfer/export/
    Returns CSV bank transfer file.
    """
    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    year = int(year_str)

    rows, _ = _get_register_rows(month, year, salary_type)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Employee ID', 'Employee Name', 'Department',
        'Bank Account Number', 'Bank IFSC Code', 'Pay Mode', 'Net Pay (INR)',
    ])
    for row in rows:
        writer.writerow([
            row['employee_id'],
            row['employee_name'],
            row['department'],
            row['bank_account'],
            row['bank_ifsc'],
            row['pay_mode'],
            f"{row['net_pay']:.2f}",
        ])

    filename = f"bank_transfer_{month}_{year}_{salary_type}.csv"
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Department Summary ───────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def department_summary(request):
    """
    GET /api/payroll/reports/department-summary/
    Returns department-wise payroll totals.
    """
    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    year = int(year_str)

    rows, run = _get_register_rows(month, year, salary_type)

    # Aggregate by department
    dept_map: dict[str, dict] = {}
    for row in rows:
        dept = row['department'] or 'Unassigned'
        if dept not in dept_map:
            dept_map[dept] = {
                'department': dept,
                'employee_count': 0,
                'total_gross': 0.0,
                'total_deductions': 0.0,
                'total_employer_contributions': 0.0,
                'total_net': 0.0,
            }
        dept_map[dept]['employee_count'] += 1
        dept_map[dept]['total_gross'] += row['gross_earnings']
        dept_map[dept]['total_deductions'] += row['total_deductions']
        dept_map[dept]['total_employer_contributions'] += row['employer_contributions']
        dept_map[dept]['total_net'] += row['net_pay']

    # Round
    for d in dept_map.values():
        for k in ('total_gross', 'total_deductions', 'total_employer_contributions', 'total_net'):
            d[k] = round(d[k], 2)

    dept_list = sorted(dept_map.values(), key=lambda x: x['department'])

    return Response({
        'success': True,
        'month': month,
        'year': year,
        'salary_type': salary_type,
        'run_id': run.id if run else None,
        'departments': dept_list,
        'totals': {
            'employee_count': sum(d['employee_count'] for d in dept_list),
            'total_gross': round(sum(d['total_gross'] for d in dept_list), 2),
            'total_deductions': round(sum(d['total_deductions'] for d in dept_list), 2),
            'total_net': round(sum(d['total_net'] for d in dept_list), 2),
        },
    })


# ─── Variance Report ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def variance_report(request):
    """
    GET /api/payroll/reports/variance/
    Returns month-over-month net pay variance per employee.
    Compares current month vs previous month.
    Query params: month, year, salary_type, threshold (default 10 — flag >10% change)
    """
    month = request.GET.get('month', '').strip()
    year_str = request.GET.get('year', '')
    salary_type = request.GET.get('salary_type', 'SALARY').upper()
    threshold = float(request.GET.get('threshold', '10'))

    if not month or not year_str:
        return Response({'success': False, 'message': 'month and year are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    year = int(year_str)

    # Current period
    curr_rows, curr_run = _get_register_rows(month, year, salary_type)
    curr_map = {r['employee_id']: r for r in curr_rows}

    # Previous period
    prev_month, prev_year = _prev_month(month, year)
    prev_rows, _ = _get_register_rows(prev_month, prev_year, salary_type)
    prev_map = {r['employee_id']: r for r in prev_rows}

    variance_rows = []
    for emp_id, curr in curr_map.items():
        prev = prev_map.get(emp_id)
        curr_net = curr['net_pay']
        prev_net = prev['net_pay'] if prev else None

        if prev_net is not None and prev_net != 0:
            change_pct = ((curr_net - prev_net) / prev_net) * 100
        elif prev_net is None:
            change_pct = None  # new employee
        else:
            change_pct = 0.0

        flagged = (
            change_pct is not None and abs(change_pct) > threshold
        )

        variance_rows.append({
            'employee_id': emp_id,
            'employee_name': curr['employee_name'],
            'department': curr['department'],
            'current_net': round(curr_net, 2),
            'previous_net': round(prev_net, 2) if prev_net is not None else None,
            'change_amount': round(curr_net - prev_net, 2) if prev_net is not None else None,
            'change_pct': round(change_pct, 2) if change_pct is not None else None,
            'flagged': flagged,
            'is_new': prev_net is None,
        })

    # Sort: flagged first, then by abs change
    variance_rows.sort(key=lambda x: (not x['flagged'], -(abs(x['change_pct'] or 0))))

    flagged_count = sum(1 for r in variance_rows if r['flagged'])
    new_count = sum(1 for r in variance_rows if r['is_new'])

    return Response({
        'success': True,
        'current_period': {'month': month, 'year': year},
        'previous_period': {'month': prev_month, 'year': prev_year},
        'salary_type': salary_type,
        'threshold_pct': threshold,
        'employee_count': len(variance_rows),
        'flagged_count': flagged_count,
        'new_employees': new_count,
        'rows': variance_rows,
    })
